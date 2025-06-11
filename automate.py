import os
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import xlrd
from difflib import SequenceMatcher
import win32com.client  # For Outlook integration
import csv

class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

class DistributorMatcherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Distributor Name Matcher")
        self.root.geometry("1100x700")

        # Create scrollable container
        self.container = ScrollableFrame(root)
        self.container.pack(fill=tk.BOTH, expand=True)

        self.distributor_data = []
        self.matches = []
        self.outlook = None

        self.create_widgets()
        self.setup_outlook()

    def setup_outlook(self):
        try:
            self.outlook = win32com.client.Dispatch("Outlook.Application")
            self.update_status("Outlook connection established")
        except Exception as e:
            messagebox.showerror("Outlook Error", f"Could not connect to Outlook: {e}")
            self.update_status("Outlook connection failed")

    def create_widgets(self):
        style = ttk.Style()
        style.configure("Treeview", rowheight=25)
        style.configure("TButton", padding=5)

        main_frame = self.container.scrollable_frame

        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=5)

        # Distributor file browse
        ttk.Label(control_frame, text="Distributor List (Excel/CSV):").grid(row=0, column=0, sticky=tk.W)
        self.distributor_entry = ttk.Entry(control_frame, width=50)
        self.distributor_entry.grid(row=0, column=1, padx=5)
        ttk.Button(control_frame, text="Browse...", command=self.browse_distributor_file).grid(row=0, column=2, padx=5)

        # Revenue folder browse
        ttk.Label(control_frame, text="Revenue Files Folder:").grid(row=1, column=0, sticky=tk.W)
        self.revenue_entry = ttk.Entry(control_frame, width=50)
        self.revenue_entry.grid(row=1, column=1, padx=5)
        ttk.Button(control_frame, text="Browse...", command=self.browse_revenue_folder).grid(row=1, column=2, padx=5)

        # Match button
        ttk.Button(control_frame, text="Find Matches", command=self.find_matches).grid(row=2, column=0, columnspan=3, pady=10)

        # Treeview for matches
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        self.tree = ttk.Treeview(tree_frame, columns=("name", "email", "status", "file", "commission", "month"), 
                                show="headings", height=12)
        for col, width in zip(["name", "email", "status", "file", "commission", "month"], [180, 200, 120, 150, 100, 100]):
            self.tree.heading(col, text=col.capitalize())
            self.tree.column(col, width=width)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        self.tree.bind('<<TreeviewSelect>>', self.show_file_preview)

        # Preview section
        preview_frame = ttk.LabelFrame(main_frame, text="File Preview", padding="10")
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        self.preview_text = tk.Text(preview_frame, wrap=tk.WORD, height=8)
        vsb_preview = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_text.yview)
        hsb_preview = ttk.Scrollbar(preview_frame, orient="horizontal", command=self.preview_text.xview)
        self.preview_text.configure(yscrollcommand=vsb_preview.set, xscrollcommand=hsb_preview.set)
        
        self.preview_text.grid(row=0, column=0, sticky="nsew")
        vsb_preview.grid(row=0, column=1, sticky="ns")
        hsb_preview.grid(row=1, column=0, sticky="ew")
        preview_frame.grid_rowconfigure(0, weight=1)
        preview_frame.grid_columnconfigure(0, weight=1)

        # Email controls frame
        email_frame = ttk.LabelFrame(main_frame, text="Email Controls", padding="10")
        email_frame.pack(fill=tk.X, pady=(10, 0))

        # Email preview
        ttk.Label(email_frame, text="Email Preview:").grid(row=0, column=0, sticky=tk.W)
        self.email_preview_text = tk.Text(email_frame, wrap=tk.WORD, height=4)
        self.email_preview_text.grid(row=1, column=0, columnspan=3, sticky="ew", pady=5)
        
        # Send buttons
        ttk.Button(email_frame, text="Send Selected", command=self.send_selected_email).grid(row=2, column=0, padx=5, pady=5)
        ttk.Button(email_frame, text="Send All Matched", command=self.send_all_matched_emails).grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(email_frame, text="Test Outlook", command=self.test_outlook).grid(row=2, column=2, padx=5, pady=5)

        self.status_var = tk.StringVar()
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, pady=(5, 0))

    def test_outlook(self):
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            messagebox.showinfo("Outlook Test", "Outlook connection successful!")
            self.update_status("Outlook test successful")
        except Exception as e:
            messagebox.showerror("Outlook Error", f"Could not connect to Outlook: {e}")
            self.update_status("Outlook test failed")

    def browse_distributor_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.csv")])
        if path:
            self.distributor_entry.delete(0, tk.END)
            self.distributor_entry.insert(0, path)
            self.load_distributor_data(path)

    def browse_revenue_folder(self):
        path = filedialog.askdirectory()
        if path:
            self.revenue_entry.delete(0, tk.END)
            self.revenue_entry.insert(0, path)
            self.update_status(f"Selected revenue folder: {path}")

    def load_distributor_data(self, path):
        try:
            ext = os.path.splitext(path)[-1].lower()
            data = []

            if ext == ".csv":
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        data.append({
                            'name': str(row.get('Distributors', '')).strip(),
                            'email': str(row.get('Distributor Email Address "TO"', '')).strip(),
                            'cc': str(row.get('Ncell Email address "CC"', '')).strip(),
                            'subject': str(row.get('Subject', '')).strip(),
                            'body': str(row.get('Body', '')).strip(),
                            'regards': str(row.get('Regards', '')).strip()
                        })
            elif ext == ".xlsx":
                wb = load_workbook(path)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    data.append({
                        'name': str(row_dict.get('Distributors', '')).strip(),
                        'email': str(row_dict.get('Distributor Email Address "TO"', '')).strip(),
                        'cc': str(row_dict.get('Ncell Email address "CC"', '')).strip(),
                        'subject': str(row_dict.get('Subject', '')).strip(),
                        'body': str(row_dict.get('Body', '')).strip(),
                        'regards': str(row_dict.get('Regards', '')).strip()
                    })
            elif ext == ".xls":
                wb = xlrd.open_workbook(path)
                sheet = wb.sheet_by_index(0)
                headers = sheet.row_values(0)
                for row_idx in range(1, sheet.nrows):
                    row = sheet.row_values(row_idx)
                    row_dict = dict(zip(headers, row))
                    data.append({
                        'name': str(row_dict.get('Distributors', '')).strip(),
                        'email': str(row_dict.get('Distributor Email Address "TO"', '')).strip(),
                        'cc': str(row_dict.get('Ncell Email address "CC"', '')).strip(),
                        'subject': str(row_dict.get('Subject', '')).strip(),
                        'body': str(row_dict.get('Body', '')).strip(),
                        'regards': str(row_dict.get('Regards', '')).strip()
                    })
            else:
                raise Exception("Unsupported file format")

            self.distributor_data = data
            self.update_status(f"Loaded {len(data)} distributors.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not load distributor file:\n{e}")
            self.update_status("Error loading file")

    def find_matches(self):
        if not self.distributor_data:
            messagebox.showwarning("Missing", "Load distributor data first")
            return

        folder = self.revenue_entry.get()
        if not folder or not os.path.exists(folder):
            messagebox.showwarning("Missing", "Select a valid revenue folder")
            return

        self.tree.delete(*self.tree.get_children())
        matched_count = 0

        for distributor in self.distributor_data:
            all_matches = []  # Store all matches for this distributor

            for file in os.listdir(folder):
                if not file.endswith(('.xlsx', '.xls', '.csv')):
                    continue
                file_path = os.path.join(folder, file)
                match_info = self.find_match_in_file(file_path, distributor['name'])
                if match_info and match_info['match_ratio'] > 0.8:  # Using our threshold
                    all_matches.append(match_info)

            if all_matches:
                matched_count += 1
                status = "✔ MATCHED"
                # Sort matches by ratio (descending)
                all_matches.sort(key=lambda x: x['match_ratio'], reverse=True)
                # Store all file paths in the tree (we'll join them with semicolons)
                file_names = ";".join([os.path.basename(m['filepath']) for m in all_matches])
                self.tree.insert("", tk.END, values=(
                    distributor['name'],
                    distributor['email'],
                    status,
                    file_names,  # Now contains all matching files
                    all_matches[0].get('commission', ''),  # Just show first match's commission
                    all_matches[0].get('month', '')       # Just show first match's month
                ), tags=('matched',))
            else:
                self.tree.insert("", tk.END, values=(distributor['name'], distributor['email'], "✖ NO MATCH", "", "", ""), tags=('unmatched',))
        
        self.tree.tag_configure('matched', background='#e6ffe6')
        self.tree.tag_configure('unmatched', background='#ffe6e6')

        self.update_status(f"Done. {matched_count}/{len(self.distributor_data)} matched.")

    def find_match_in_file(self, path, target_name):
        ext = os.path.splitext(path)[-1].lower()

        try:
            headers = []
            rows = []

            if ext == ".xlsx":
                wb = load_workbook(path, read_only=True)
                sheet = wb.active
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    rows.append(row)
            elif ext == ".xls":
                wb = xlrd.open_workbook(path)
                sheet = wb.sheet_by_index(0)
                headers = sheet.row_values(0)
                for i in range(1, sheet.nrows):
                    rows.append(sheet.row_values(i))
            elif ext == ".csv":
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader)
                    for row in reader:
                        rows.append(row)
            else:
                return None

            possible_keys = ["Distributors", "Distributor' Name", "Distributor Name"]
            name_col = next((headers.index(k) for k in possible_keys if k in headers), None)

            if name_col is None:
                return None

            best = None
            best_ratio = 0

            for row in rows:
                name_in_file = str(row[name_col]).strip()
                ratio = SequenceMatcher(None, name_in_file.lower(), target_name.lower()).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    row_data = dict(zip(headers, row))
                    best = {
                        'filepath': path,
                        'match_ratio': ratio,
                        'commission': row_data.get('Package Number', '') or row_data.get('Commission', ''),
                        'month': row_data.get('Ecare Month', '') or row_data.get('Month', ''),
                        'headers': headers,
                        'rows': rows
                    }

            if best:
                best['match_ratio'] = best_ratio
            return best
        except Exception as e:
            self.update_status(f"Error reading {path}: {e}")
            return None

    def show_file_preview(self, event):
        selected_item = self.tree.focus()
        if not selected_item:
            return
            
        item_values = self.tree.item(selected_item, 'values')
        if len(item_values) < 4 or not item_values[3]:
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(tk.END, "No file selected or no match found")
            self.email_preview_text.delete(1.0, tk.END)
            return
            
        file_names = item_values[3].split(';')
        folder = self.revenue_entry.get()
        
        distributor_name = item_values[0]
        distributor = next((d for d in self.distributor_data if d['name'] == distributor_name), None)
        
        if not distributor:
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(tk.END, "Could not find distributor details")
            self.email_preview_text.delete(1.0, tk.END)
            return
            
        # Build complete email body with regards
        complete_body = distributor['body']
        if distributor.get('regards'):
            complete_body += f"\n\nRegards,\n{distributor['regards']}"
        
        self.email_preview_text.delete(1.0, tk.END)
        self.email_preview_text.insert(tk.END, f"To: {distributor['email']}\n")
        self.email_preview_text.insert(tk.END, f"Cc: {distributor['cc']}\n")
        self.email_preview_text.insert(tk.END, f"Subject: {distributor['subject']}\n")
        self.email_preview_text.insert(tk.END, f"\nBody:\n{complete_body}")

        self.preview_text.delete(1.0, tk.END)
        self.preview_text.insert(tk.END, f"Found {len(file_names)} matching files:\n\n")
        
        for i, file_name in enumerate(file_names, 1):
            file_path = os.path.join(folder, file_name.strip())
            if not os.path.exists(file_path):
                self.preview_text.insert(tk.END, f"{i}. {file_name} (File not found)\n\n")
                continue
                
            match_info = self.find_match_in_file(file_path, distributor_name)
            
            if not match_info:
                self.preview_text.insert(tk.END, f"{i}. {file_name} (Could not load match details)\n\n")
                continue
                
            self.preview_text.insert(tk.END, f"{i}. {file_name} (Match ratio: {match_info['match_ratio']:.2f})\n")
            
            headers = match_info.get('headers', [])
            self.preview_text.insert(tk.END, "Headers:\n")
            self.preview_text.insert(tk.END, "\t".join(str(h) for h in headers) + "\n")
            
            rows = match_info.get('rows', [])
            self.preview_text.insert(tk.END, f"First {min(3, len(rows))} rows:\n")
            for row in rows[:3]:
                self.preview_text.insert(tk.END, "\t".join(str(cell) for cell in row) + "\n")
                
            if len(rows) > 3:
                self.preview_text.insert(tk.END, f"... and {len(rows)-3} more rows\n")
            
            self.preview_text.insert(tk.END, "\n")

    def send_selected_email(self):
        if not self.outlook:
            messagebox.showerror("Error", "Outlook connection not established")
            return

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a distributor from the list")
            return
            
        item_values = self.tree.item(selected_item, 'values')
        if len(item_values) < 4 or not item_values[3] or "MATCHED" not in item_values[2]:
            messagebox.showwarning("No Match", "Selected distributor has no matched file")
            return
            
        distributor_name = item_values[0]
        distributor = next((d for d in self.distributor_data if d['name'] == distributor_name), None)
        
        if not distributor:
            messagebox.showerror("Error", "Could not find distributor details")
            return
            
        file_names = item_values[3].split(';')
        folder = self.revenue_entry.get()
        
        try:
            mail = self.outlook.CreateItem(0)
            mail.To = distributor['email']
            
            if distributor['cc'] and "@" in distributor['cc']:
                mail.CC = distributor['cc']
            
            mail.Subject = distributor['subject']
            
            # Build complete body with regards
            complete_body = distributor['body']
            if distributor.get('regards'):
                complete_body += f"\n\nRegards,\n{distributor['regards']}"
            mail.Body = complete_body
            
            # Add all attachments
            attachments_added = 0
            for file_name in file_names:
                file_path = os.path.join(folder, file_name.strip())
                if os.path.exists(file_path):
                    mail.Attachments.Add(file_path)
                    attachments_added += 1
                else:
                    messagebox.showwarning("File Missing", f"Could not find file: {file_path}")
            
            if attachments_added == 0:
                messagebox.showerror("Error", "No valid attachments found")
                return
            
            try:
                mail.Recipients.ResolveAll()
            except Exception as resolve_error:
                unresolved = [r.Name for r in mail.Recipients if not r.Resolved]
                messagebox.showerror("Recipient Error",
                                   f"Could not resolve recipients: {', '.join(unresolved)}")
                return
                
            mail.Display(True)
            
            self.update_status(f"Email prepared for {distributor_name} with {attachments_added} attachments. Please review and send.")
            
        except Exception as e:
            messagebox.showerror("Email Error", f"Could not create email: {str(e)}")
            self.update_status(f"Email creation failed: {str(e)}")

    def send_all_matched_emails(self):
        if not self.outlook:
            messagebox.showerror("Error", "Outlook connection not established")
            return

        matched_items = [self.tree.item(item) for item in self.tree.get_children() 
                       if "MATCHED" in self.tree.item(item, 'values')[2]]
        
        if not matched_items:
            messagebox.showinfo("No Matches", "No matched distributors found")
            return
            
        confirm = messagebox.askyesno("Confirm", 
                                    f"Send emails to all {len(matched_items)} matched distributors?")
        if not confirm:
            return
            
        folder = self.revenue_entry.get()
        success_count = 0
        failed_count = 0
        failed_distributors = []
        
        progress = tk.Toplevel(self.root)
        progress.title("Sending Emails")
        progress.geometry("400x150")
        
        tk.Label(progress, text="Sending emails...").pack(pady=5)
        progress_bar = ttk.Progressbar(progress, orient="horizontal", length=350, mode="determinate")
        progress_bar.pack(pady=5)
        progress_bar["maximum"] = len(matched_items)
        status_label = tk.Label(progress, text="")
        status_label.pack()
        details_label = tk.Label(progress, text="", wraplength=350)
        details_label.pack()
        
        progress.update()
        
        for item in matched_items:
            values = item['values']
            distributor_name = values[0]
            status_label.config(text=f"Processing: {distributor_name}")
            details_label.config(text="")
            progress.update()
            
            distributor = next((d for d in self.distributor_data if d['name'] == distributor_name), None)
            
            if not distributor:
                details_label.config(text="No distributor data found")
                failed_count += 1
                failed_distributors.append(f"{distributor_name} - No data")
                progress_bar["value"] += 1
                continue
                
            if not self.is_valid_email(distributor['email']):
                details_label.config(text=f"Invalid email: {distributor['email']}")
                failed_count += 1
                failed_distributors.append(f"{distributor_name} - Bad email")
                progress_bar["value"] += 1
                continue
                
            file_names = values[3].split(';')
            attachments_added = 0
            
            try:
                mail = self.outlook.CreateItem(0)
                mail.To = distributor['email']
                
                if distributor.get('cc'):
                    cleaned_cc = self.clean_email_list(distributor['cc'])
                    if cleaned_cc:
                        mail.CC = cleaned_cc
                
                mail.Subject = distributor['subject']
                
                # Build complete body with regards
                complete_body = distributor['body']
                if distributor.get('regards'):
                    complete_body += f"\n\nRegards,\n{distributor['regards']}"
                mail.Body = complete_body
                
                # Add all attachments
                for file_name in file_names:
                    file_path = os.path.join(folder, file_name.strip())
                    if os.path.exists(file_path):
                        mail.Attachments.Add(file_path)
                        attachments_added += 1
                
                if attachments_added == 0:
                    raise ValueError("No valid attachments found")
                
                unresolved = []
                try:
                    mail.Recipients.ResolveAll()
                    unresolved = [r.Name for r in mail.Recipients if not r.Resolved]
                except Exception as resolve_error:
                    unresolved = [r.Name for r in mail.Recipients]
                
                if unresolved:
                    raise ValueError(f"Unresolved recipients: {', '.join(unresolved)}")
                
                mail.Send()
                success_count += 1
                details_label.config(text=f"Sent with {attachments_added} attachments")
                
            except Exception as e:
                error_msg = str(e)
                details_label.config(text=error_msg)
                failed_count += 1
                failed_distributors.append(f"{distributor_name} - {error_msg}")
                
            progress_bar["value"] += 1
            time.sleep(1)
            progress.update()
            
        progress.destroy()
        
        result_message = f"Completed: {success_count} sent, {failed_count} failed"
        if failed_distributors:
            result_message += "\n\nFailed items:\n" + "\n".join(failed_distributors)
        
        messagebox.showinfo("Results", result_message)
        self.update_status(f"Email sending complete. {success_count} sent, {failed_count} failed")

    def is_valid_email(self, email):
        """Basic email validation"""
        if not email or not isinstance(email, str):
            return False
        return '@' in email and '.' in email.split('@')[-1]

    def clean_email_list(self, emails):
        """Clean and validate a list of emails (for CC field)"""
        if not emails:
            return ""
        valid_emails = []
        for email in emails.split(';'):
            email = email.strip()
            if self.is_valid_email(email):
                valid_emails.append(email)
        return ";".join(valid_emails)

    def update_status(self, msg):
        self.status_var.set(msg)
        self.root.update_idletasks()

if __name__ == "__main__":
    root = tk.Tk()
    app = DistributorMatcherApp(root)
    root.mainloop()